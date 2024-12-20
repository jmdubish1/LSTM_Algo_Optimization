import numpy as np
import pandas as pd
import openpyxl
import os
from datetime import datetime, timedelta
from openpyxl.drawing.image import Image
import nn_tools.loss_functions as lf
from nn_tools.custom_callbacks_layers import TemperatureScalingLayer
import keras
import pickle
import re

from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from nn_tools.process_handler import ProcessHandler


class SaveHandler:
    def __init__(self, process_handler: "ProcessHandler"):
        self.ph = process_handler
        self.ph.save_handler = self
        self.save_file = None
        self.trade_metrics = None

        self.test_date = None
        self.end_date = None
        self.train_date = None

        self.model_summary = None

        self.param_folder = ''
        self.data_folder = ''
        self.model_folder = ''
        self.main_train_path = ''
        self.previous_model_path = ''
        self.model_save_path = ''

    def check_create_model_folder(self):
        self.param_folder = \
            (f'{self.ph.setup_params.trade_dat_loc}\\{self.ph.setup_params.security}\\'
             f'{self.ph.setup_params.time_frame_test}\\'
             f'{self.ph.setup_params.time_frame_test}_test_{self.ph.setup_params.time_len}\\'
             f'{self.ph.setup_params.model_type}\\{self.ph.side}')

        self.data_folder = f'{self.param_folder}\\Data'
        self.model_folder = f'{self.param_folder}\\Models'
        self.main_train_path = f'{self.model_folder}\\{self.ph.side}\\param_{self.ph.paramset_id}_main_model'

        for folder in [self.param_folder, self.data_folder, self.model_folder]:
            os.makedirs(folder, exist_ok=True)

    def set_model_train_paths(self):
        test_date = self.ph.trade_data.curr_test_date
        self.check_create_model_folder()

        previous_test_date = pd.to_datetime(test_date) - timedelta(days=self.ph.setup_params.test_period_days)
        previous_test_date = previous_test_date.strftime(format='%Y-%m-%d')[:11]
        self.test_date = test_date
        test_date = test_date.strftime(format='%Y-%m-%d')[:11]

        self.model_save_path = \
            f'{self.model_folder}\\{self.ph.side}_{self.ph.paramset_id}\\{test_date}_model'
        self.previous_model_path = \
            f'{self.model_folder}\\{self.ph.side}_{self.ph.paramset_id}\\{previous_test_date}_model'

    def save_model(self, i):
        print(f'Model Saved: {self.model_save_path}')
        self.ph.lstm_model.model.save(f'{self.model_save_path}\\model.keras')
        if i == 0:
            self.ph.lstm_model.model.save(f'{self.main_train_path}\\model.keras')

    def load_model(self, model_path):
        if self.ph.setup_params.model_type == 'classification_lstm':
            penalty_cat_crossentropy = lf.penalized_categorical_crossentropy(self.ph.lstm_model.penalty_matrix)

            self.ph.lstm_model.model = (
                keras.models.load_model(f'{model_path}',
                                        custom_objects={'penalized_crossentropy_loss': penalty_cat_crossentropy}))
            self.ph.lstm_model.model.load_weights(f'{model_path}')

    def load_prior_test_date_model(self):
        model_path = f'{self.previous_model_path}\\model.keras'
        last_test_date = (pd.to_datetime(self.test_date, format='%Y-%m-%d') -
                          timedelta(days=self.ph.setup_params.test_period_days)).strftime(format='%Y-%m-%d')
        self.load_model(model_path)
        print(f'Loading Prior Week Model: {str(last_test_date)}')

    def load_current_test_date_model(self):
        model_path = f'{self.model_save_path}\\model.keras'
        self.load_model(model_path)
        print(f'Loading Current Week Model: {self.ph.side}: {self.ph.paramset_id}: {self.test_date}')

    def save_scalers(self):
        scalers = {
            'y_pnl_scaler': self.ph.lstm_data.y_pnl_scaler,
            'intra_scaler': self.ph.lstm_data.intra_scaler,
            'daily_scaler': self.ph.lstm_data.daily_scaler
        }
        os.makedirs(os.path.dirname(f'{self.model_save_path}\\'), exist_ok=True)
        for key, val in scalers.items():
            with open(f'{self.model_save_path}\\{key}.pkl', 'wb') as f:
                pickle.dump(val, f)
            print(f'Saved {key} Scaler\n')

    def load_scalers(self, retrain=False):
        curr_scalers_exist = os.path.exists(f'{self.model_save_path}\\intra_scaler.pkl')
        prev_scalers_exist = os.path.exists(f'{self.previous_model_path}\\intra_scaler.pkl')

        if (not retrain or not curr_scalers_exist) and prev_scalers_exist:
            with open(f'{self.previous_model_path}\\y_pnl_scaler.pkl', 'rb') as f:
                self.ph.lstm_data.y_pnl_scaler = pickle.load(f)
            with open(f'{self.previous_model_path}\\intra_scaler.pkl', 'rb') as f:
                self.ph.lstm_data.intra_scaler = pickle.load(f)
            with open(f'{self.previous_model_path}\\daily_scaler.pkl', 'rb') as f:
                self.ph.lstm_data.daily_scaler = pickle.load(f)
                print(f'...Loaded Previous Scalers: \n'
                      f'{self.previous_model_path}')

        elif curr_scalers_exist:
            with open(f'{self.model_save_path}\\y_pnl_scaler.pkl', 'rb') as f:
                self.ph.lstm_data.y_pnl_scaler = pickle.load(f)
            with open(f'{self.model_save_path}\\intra_scaler.pkl', 'rb') as f:
                self.ph.lstm_data.intra_scaler = pickle.load(f)
            with open(f'{self.model_save_path}\\daily_scaler.pkl', 'rb') as f:
                self.ph.lstm_data.daily_scaler = pickle.load(f)
                print(f'Loaded Previous Scalers: \n'
                      f'...{self.model_save_path}')
        else:
            pass

    def save_all_prediction_data(self, test_date, model_dfs, include_small_tf):
        self.save_metrics(self.ph.side, self.ph.paramset_id, test_date, model_dfs[1], 'Model')
        self.save_metrics(self.ph.side, self.ph.paramset_id, test_date, model_dfs[0], 'PnL', include_small_tf)
        if self.ph.train_modeltf and not self.ph.retraintf:
            self.save_plot_to_excel(self.ph.side)

    def save_metrics(self, side, param, test_date, dfs, sheet_name, stack_row=False, include_small_tf=False):
        os.makedirs(f'{self.data_folder}\\{side}_{param}', exist_ok=True)
        if include_small_tf:
            self.save_file = f'{self.data_folder}\\{side}_{param}\\predictions_{side}_{param}_{test_date}_small.xlsx'
        else:
            self.save_file = f'{self.data_folder}\\{side}_{param}\\predictions_{side}_{param}_{test_date}.xlsx'
        sheet_name = f'{side}_{sheet_name}'

        if os.path.exists(self.save_file):
            # Load the existing workbook
            book = openpyxl.load_workbook(self.save_file)
            if not book.sheetnames:
                book.create_sheet(sheet_name, -1)
                book.active.title = sheet_name

            with pd.ExcelWriter(self.save_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                start_positions = get_excel_sheet_df_positions(dfs, stack_row)
                write_metrics_to_excel(writer, dfs, sheet_name, start_positions)

        else:
            # Create a new Excel file
            create_new_excel_file(self.save_file, sheet_name)
            with pd.ExcelWriter(self.save_file, engine='openpyxl') as writer:
                start_positions = get_excel_sheet_df_positions(dfs, stack_row)
                write_metrics_to_excel(writer, dfs, sheet_name, start_positions)

    def save_plot_to_excel(self, side):
        file_exists = os.path.exists(self.save_file)
        if not file_exists:
            wb = openpyxl.Workbook()
        else:
            wb = openpyxl.load_workbook(self.save_file)

        # Select a sheet or create a new one
        sheet_name = f'{side}_LR_Curve'
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)

        img_loc = f'{self.data_folder}\\temp_img.png'
        if self.ph.lstm_model.model_plot:
            self.ph.lstm_model.model_plot.fig.savefig(img_loc)
        img = Image(img_loc)

        plot_loc_excel = 'F2'
        ws.add_image(img, plot_loc_excel)

        wb.save(self.save_file)

        if os.path.exists(img_loc):
            os.remove(img_loc)

    def save_opt_thres_temp(self, side, param, test_date, opt_threshold, opt_temp):
        temp_df = pd.DataFrame([[side, param, test_date, opt_threshold, opt_temp]],
                               columns=['side', 'paramset_id', 'test_date', 'opt_threshold', 'opt_temp'])

        opt_thres_file = f'{self.param_folder}\\best_thresholds.xlsx'

        if os.path.exists(opt_thres_file):
            try:
                existing_data = pd.read_excel(opt_thres_file, engine='openpyxl')
            except Exception as e:
                print(f"Error loading the existing Excel file: {e}")
                return

            if param in existing_data['paramset_id'].values:
                existing_data.loc[existing_data['paramset_id'] == param, :] = temp_df.values[0]
            else:
                existing_data = pd.concat([existing_data, temp_df], ignore_index=True)

            existing_data.to_excel(opt_thres_file, index=False, engine='openpyxl')

        else:
            temp_df.to_excel(opt_thres_file, index=False, engine='openpyxl')


def write_metrics_to_excel(writer, dfs, sheet_name, start_positions):
    for df, (startrow, startcol) in zip(dfs, start_positions):
        df.to_excel(writer, sheet_name=sheet_name,
                    startrow=startrow, startcol=startcol)


def get_excel_sheet_df_positions(dfs, stack_row):
    start_positions = [(0, 0)]

    if len(dfs) > 1:
        if stack_row:
            start_row = len(dfs[0])
            for df in dfs[1:]:
                start_row += 2
                start_positions.append((start_row, 0))
                start_row += len(df)
        else:
            start_row = 0
            start_col = len(dfs[0].columns) + 2
            for df in dfs[1:]:
                start_positions.append((start_row, start_col))
                start_row += len(df) + 2

    return start_positions


def create_new_excel_file(file_path, sheet_name):
    if not os.path.exists(file_path):
        df = pd.DataFrame()
        df.to_excel(file_path, sheet_name=sheet_name)


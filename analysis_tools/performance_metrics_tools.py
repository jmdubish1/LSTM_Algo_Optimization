import numpy as np
import pandas as pd
import os
import re
import data_tools.math_tools as mt
from openpyxl import load_workbook

setup_dict1 = {
    'strategy': 'Double_Candle',
    'model_type': 'LSTM',
    'security': 'NQ',
    'side': 'Bear',
    'time_frame': '15min',
    'time_length': '20years',
    'trade_dat_loc': r'C:\Users\jmdub\Documents\Trading\Futures\Strategy Info\Double_Candles\ATR',
    'start_train_date': '2013-04-01',
    'final_test_date': '2024-04-01',
    'start_hour': 5,
    'start_minute': 30,
    'test_period_days': 7*13,
    'years_to_train': 3,
    'total_param_sets': 289
}

performance_targets1 = {
    'vars': [.99, .999, .9999],
    'max_drawdown': 3000.0,
    'max_draw_months': 2.0,
    'profit_factor': 2.0,
    'expect_ratio': 0.5,
    'sharp_ratio': 2.0,
    'sortino_ratio': 0.8
}


class AllocationMetrics:
    def __init__(self, setup_dict):
        self.setup_dict = setup_dict
        self.trade_dat_loc = setup_dict['trade_dat_loc']
        self.strat_main = (f'{self.trade_dat_loc}\\{self.setup_dict["security"]}\\{self.setup_dict["time_frame"]}\\'
                           f'{self.setup_dict["time_frame"]}_test_{self.setup_dict["time_length"]}')
        self.model_dat_folder = f'{self.strat_main}\\{self.setup_dict["side"]}\\Data'
        self.trades_file = f'{self.strat_main}\\NQ_15min_Double_Candle_289_trades.feather'
        self.params_file = f'{self.strat_main}\\NQ_15min_Double_Candle_289_params.feather'
        self.side = setup_dict['side']
        self.models_to_test = []
        self.params_to_test = []

        self.working_param_dfs = []
        self.agged_params = None

    def get_models_to_test(self):
        param_id = re.compile(r'.*_(\d+)$')

        for item in os.listdir(self.model_dat_folder):
            param_match = param_id.match(item)
            item_path = os.path.join(self.model_dat_folder, item)
            if os.path.isdir(item_path) and param_match:
                self.models_to_test.append(item)
                self.params_to_test.append(int(param_match.group(1)))

        print(self.models_to_test)
        print(self.params_to_test)

    def load_concat_datasets(self, param_id):
        ind = self.params_to_test.index(param_id)
        dfs = []
        file_folder = f'{self.model_dat_folder}\\{self.models_to_test[ind]}'
        for item in os.listdir(file_folder):
            item_path = os.path.join(file_folder, item)
            df = pd.read_excel(item_path, sheet_name=f'{self.side}_WL', usecols='B:N')
            dfs.append(df)

        combined_df = pd.concat(dfs)
        combined_df['DateTime'] = pd.to_datetime(combined_df['DateTime'], format='%Y-%m-%d %H:%M:%S')
        combined_df = combined_df.sort_values(by='DateTime')
        combined_df.reset_index(inplace=True, drop=True)
        combined_df = arrange_model_columns(combined_df)

        return combined_df


class ModelPerformanceMetrics:
    def __init__(self, model_data, param_id):
        self.model_data = model_data.copy(deep=True)
        self.analysis_df = None
        self.param_id = param_id

    def concat_dfs(self):
        dfs = []
        for col in ['Algo', 'Pred', 'Two_Dir_Pred']:
            temp_df = self.model_data[['DateTime', f'{self.param_id}_{col}_PnL']].copy(deep=True)
            temp_df['strat'] = f'{self.param_id}_{col}'
            temp_df.rename(columns={f'{self.param_id}_{col}_PnL': f'{self.param_id}_PnL'}, inplace=True)
            dfs.append(temp_df)

        df = pd.concat(dfs)
        df = df.sort_values(by='DateTime')
        df.reset_index(inplace=True, drop=True)

        self.analysis_df = df

    def build_allocation_columns(self):
        self.analysis_df[f'{self.param_id}_PnL_Total'] = self.analysis_df[f'{self.param_id}_PnL'].cumsum()
        self.analysis_df[f'{self.param_id}_MaxDraw'] = (
            mt.calculate_max_drawdown(self.analysis_df[f'{self.param_id}_PnL_Total']))

        self.analysis_df = self.analysis_df[['DateTime', 'strat', f'{self.param_id}_PnL',
                                             f'{self.param_id}_PnL_Total', f'{self.param_id}_MaxDraw']]
        self.analysis_df['DateTime'] = pd.to_datetime(self.analysis_df['DateTime'],
                                                      format='%Y-%m-%d %HH:%MM:%SS', errors='coerce')
        self.analysis_df = self.analysis_df.dropna(subset='DateTime')

    def set_allo_metrics(self, save_path):
        wb = load_workbook(save_path)
        if f'{self.param_id}_Allo_metrics' in wb.sheetnames:
            del wb[f'{self.param_id}_Allo_metrics']

        allo_sheet = wb.create_sheet(title=f'{self.param_id}_Allo_metrics')
        p = self.param_id

        # Allocation settings
        allo_sheet['A1'] = 'Allocation Adjustment'
        allo_sheet['A2'] = 'Security Value'
        allo_sheet['B2'] = '2'
        allo_sheet['A4'] = 'Strategy'
        allo_sheet['B4'] = 'Allocation'
        allo_sheet['B5'] = 'Total'
        allo_sheet['A6'] = f'{p}_Algo'
        allo_sheet['A7'] = f'{p}_Pred'
        allo_sheet['A8'] = f'{p}_Two_Dir_Pred'
        allo_sheet['B6'] = 0
        allo_sheet['B7'] = 1
        allo_sheet['B8'] = 1

        # Daily Data
        allo_sheet['E1'] = 'Daily Data'

        allo_sheet['E2'] = 'Winning Days'
        allo_sheet['E3'] = f'=COUNTIF(\'{p}_agged\'!M:M, ">"&0)'
        allo_sheet['F2'] = 'Losing Days'
        allo_sheet['F3'] = f'=COUNTIF(\'{p}_agged\'!M:M, "<"&0)'
        allo_sheet['G2'] = '% Winning'
        allo_sheet['G3'] = '=E3/SUM(E3:F3)'
        allo_sheet['H2'] = 'Worst Day'
        allo_sheet['H3'] = f"=MIN('{p}_agged'!M:M)"
        allo_sheet['I2'] = 'Best Day'
        allo_sheet['I3'] = f"=MAX('{p}_agged'!M:M)"

        allo_sheet['E5'] = 'Avg Win'
        allo_sheet['E6'] = f'=AVERAGEIF(\'{p}_agged\'!M:M,">"&0)'
        allo_sheet['F5'] = 'Avg Loss'
        allo_sheet['F6'] = f'=AVERAGEIF(\'{p}_agged\'!M:M,"<"&0)'
        allo_sheet['G5'] = 'Avg Diff'
        allo_sheet['G6'] = '=E6+F6'
        allo_sheet['H5'] = 'Expected Daily'
        allo_sheet['H6'] = '=E6*G3+(1-G3)*F6'

        allo_sheet['E8'] = 'Avg # of Trades'
        allo_sheet['E9'] = f'=AVERAGE(\'{p}_agged\'!N:N)'
        allo_sheet['F8'] = 'Avg # of Winning Trades'
        allo_sheet['F9'] = '=G13*E9'
        allo_sheet['G8'] = 'Avg # of Losing Trades'
        allo_sheet['G9'] = '=(1-G13)*E9'
        allo_sheet['H8'] = 'Daily Commission'
        allo_sheet['H9'] = '=E22/SUM(E3:F3)'

        # Trade Data
        allo_sheet['E11'] = 'Daily Data'
        allo_sheet['E12'] = 'Winning Trades'
        allo_sheet['E13'] = f'=COUNTIF(\'{p}_agged\'!G:G,">"&0)'
        allo_sheet['F12'] = 'Losing Trades'
        allo_sheet['F13'] = f'=COUNTIF(\'{p}_agged\'!G:G,"<"&0)'
        allo_sheet['G12'] = '% Winning'
        allo_sheet['G13'] = '=E13/SUM(E13:F13)'
        allo_sheet['H12'] = 'Days Below $3K'
        allo_sheet['H13'] = f'=COUNTIF(\'{p}_agged\'!M:M, "<="&-3000)'
        allo_sheet['I12'] = 'Days Above $5K'
        allo_sheet['I13'] = f'=COUNTIF(\'{p}_agged\'!M:M, ">="&5000)'

        allo_sheet['E15'] = 'Avg Win'
        allo_sheet['E16'] = f'=AVERAGEIF(\'{p}_agged\'!G:G,">"&0)'
        allo_sheet['F15'] = 'Avg Loss'
        allo_sheet['F16'] = f'=AVERAGEIF(\'{p}_agged\'!G:G,"<"&0)'
        allo_sheet['G15'] = 'Avg Diff'
        allo_sheet['G16'] = '=E16+F16'
        allo_sheet['H15'] = 'Total PnL - Wins'
        allo_sheet['H16'] = f'=SUMIF(\'{p}_agged\'!G:G, ">"&0)'
        allo_sheet['I15'] = 'Total PnL - Losses'
        allo_sheet['I16'] = f'=SUMIF(\'{p}_agged\'!G:G, "<"&0)'

        # Overall Performance
        allo_sheet['E18'] = 'Overall Performance'
        allo_sheet['E19'] = 'Max DrawDown'
        allo_sheet['E20'] = f'=MAX(\'{p}_agged\'!I:I)'
        allo_sheet['F19'] = 'Net Profit'
        allo_sheet['F20'] = f'=LOOKUP(2,1/(\'{p}_agged\'!H:H<>""),\'{p}_agged\'!H:H)'
        allo_sheet['G19'] = 'Net To DrawDown'
        allo_sheet['G20'] = '=F20/E20'

        allo_sheet['E21'] = 'Total Commission'
        allo_sheet['E22'] = f'=SUM(\'{p}_agged\'!F:F)*4'
        allo_sheet['F21'] = 'Net of Commission'
        allo_sheet['F22'] = '=F20-E22'
        allo_sheet['G21'] = 'Net to DrawDown'
        allo_sheet['G22'] = '=F22/E20'

        # VaR
        allo_sheet['F24'] = 'String of Losses'
        allo_sheet['G24'] = 'Notional DrawDown'
        allo_sheet['E25'] = '1 in 252 Days'
        allo_sheet['E26'] = '1 in 60 Days'
        allo_sheet['E27'] = '1 in 30 Days'
        allo_sheet['F25'] = '=-LOG(252/(1-$G$3)-1)/LOG(1-$G$3)'
        allo_sheet['F26'] = '=-LOG(60/(1-$G$3)-1)/LOG(1-$G$3)'
        allo_sheet['F27'] = '=-LOG(30/(1-$G$3)-1)/LOG(1-$G$3)'
        allo_sheet['G25'] = '=F25*F$6'
        allo_sheet['G26'] = '=F26*F$6'
        allo_sheet['G27'] = '=F27*F$6'

        # Ratios
        allo_sheet['H18'] = 'Ratios'
        allo_sheet['H19'] = 'Kelly Criterion'
        allo_sheet['H10'] = 'Sharpe'
        allo_sheet['H21'] = 'Sortino'
        allo_sheet['H22'] = 'Expectation Ratio'
        allo_sheet['H23'] = 'Profit Factor'
        allo_sheet['I19'] = '=G13-(1-G13)/ABS(E16/F16)'
        allo_sheet['I20'] = 'Need Margin'
        allo_sheet['I21'] = 'Need Margin'
        allo_sheet['I22'] = '=-H6/F6'
        allo_sheet['I23'] = '=H16/-I16'

        wb.save(save_path)

    def agged_excel_work(self, save_path):
        wb = load_workbook(save_path)
        p = self.param_id
        agg_sheet = wb[f'{p}_agged']

        agg_sheet['F1'] = f'{p}_Contracts'
        agg_sheet['G1'] = f'{p}_Adj_PnL'
        agg_sheet['H1'] = f'{p}_Adj_PnL_Total'
        agg_sheet['H2'] = '=G2'
        agg_sheet['I1'] = f'{p}_Adj_MaxDraw'
        agg_sheet['L1'] = f'{p}_Daily'
        agg_sheet['J1'] = f'{p}_Date'
        agg_sheet['M1'] = f'{p}_Daily_PnL'
        agg_sheet['N1'] = f'{p}_Count'

        for i in range(2, agg_sheet.max_row + 1):
            agg_sheet[f'F{i}'] = f'=VLOOKUP(B{i},{p}_Allo_metrics!$A$6:$B$8, 2, FALSE)'
            agg_sheet[f'G{i}'] = f'=C{i}*F{i}*{p}_Allo_metrics!$B$2'
            agg_sheet[f'H{i+1}'] = f'=H{i}+G{i+1}'
            agg_sheet[f'I{i}'] = f'=IF(H{i}<0,-H{i},-MIN((H{i}-MAX($H$2:H{i}))))'
            agg_sheet[f'J{i}'] = f'=TEXT(A{i},"yyyy-mm-dd")'

        uniq_dates = np.unique(self.analysis_df['DateTime'].dt.strftime('%Y-%m-%d'))
        for i in range(2, len(uniq_dates) + 1):
            agg_sheet[f'L{i}'] = uniq_dates[i-2]
            agg_sheet[f'M{i}'] = f'=SUMIF(J:J,"="&L{i},G:G)'
            agg_sheet[f'N{i}'] = f'=COUNTIF(J:J,"="&L{i})'

        wb.save(save_path)


def arrange_model_columns(df):
    priority_columns = ['DateTime', 'Close', 'Algo_wl', 'Pred_wl']
    existing_priority_columns = [col for col in priority_columns if col in df.columns]
    remaining_columns = [col for col in df.columns if col not in existing_priority_columns]
    df = df[existing_priority_columns + remaining_columns]

    return df


def rename_cols(df, param_id):
    """Should be a method"""
    if 'Two_Dir_Pred_Pnl_Total' in df.columns:
        df.drop(columns=['Two_Dir_Pred_Pnl_Total'], inplace=True)
        column_to_move = df.pop('Two_Dir_Pred_PnL_Total')
        df.insert(len(df.columns) - 1, 'Two_Dir_Pred_PnL_Total', column_to_move)

    col_dict = {}
    cols_to_rename = df.columns
    for col in cols_to_rename:
        if col not in ['DateTime', 'Close']:
            col_dict[f'{col}'] = f'{param_id}_{col}'

    df.rename(columns=col_dict, inplace=True)

    return df


def create_rolling_sum(df):
    pnl_cols = ["Algo", "Pred", "Two_Dir_Pred"]
    for col in pnl_cols:
        df[f'{col}_PnL_Total'] = df[f'{col}_PnL'].cumsum()
        df[f'{col}_MaxDraw'] = mt.calculate_max_drawdown(df[f'{col}_PnL_Total'])

    return df


def create_rolling_sum_agged(df, param):
    df[f'{param}_PnL_Total'] = df[f'PnL'].cumsum()

    return df


def add_sheet_excel(df, sheet_name, file_path):
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    else:
        combined_df.to_excel(file_path, sheet_name=sheet_name, index=False)


perfmet = AllocationMetrics(setup_dict1)
perfmet.get_models_to_test()
for param in perfmet.params_to_test:
    print(f'Analyzing: {param}')
    combined_df = perfmet.load_concat_datasets(param)
    combined_df = create_rolling_sum(combined_df)
    combined_df = rename_cols(combined_df, param)
    save_path = f'{perfmet.model_dat_folder}\\Agg_Data\\agg_data_{perfmet.side}_{param}.xlsx'
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    add_sheet_excel(combined_df, f'{param}_Full_data', save_path)

    model_met = ModelPerformanceMetrics(combined_df, param)
    model_met.concat_dfs()
    model_met.build_allocation_columns()
    add_sheet_excel(model_met.analysis_df, f'{param}_agged', save_path)
    model_met.set_allo_metrics(save_path)
    model_met.agged_excel_work(save_path)
